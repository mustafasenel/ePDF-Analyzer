"""Export manager for converting data to various formats"""

import json
import pandas as pd
from pathlib import Path
from typing import Dict, List, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from config.settings import settings
from app.utils.helpers import sanitize_sheet_name


class ExportError(Exception):
    """Export operation error"""
    pass


class ExportManager:
    """
    Manager for exporting data to different formats
    """
    
    def export_to_json(
        self,
        data: Dict[str, Any],
        output_path: Optional[str | Path] = None,
        pretty: bool = True
    ) -> str | Dict:
        """
        Export data to JSON format
        
        Args:
            data: Data to export
            output_path: Output file path (None to return as string)
            pretty: Pretty print JSON
            
        Returns:
            JSON string if output_path is None, otherwise file path
        """
        indent = 2 if pretty else None
        
        if output_path:
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            with output_path.open('w', encoding='utf-8') as f:
                json.dump(data, f, indent=indent, ensure_ascii=False)
            
            return str(output_path)
        else:
            return json.dumps(data, indent=indent, ensure_ascii=False)
    
    def export_to_excel(
        self,
        tables: Dict[int, List[pd.DataFrame]],
        output_path: str | Path,
        include_text: bool = False,
        text_data: Optional[Dict[int, str]] = None,
        add_styling: bool = True
    ) -> str:
        """
        Export tables to Excel file
        
        Args:
            tables: Dictionary mapping page numbers to list of DataFrames
            output_path: Output Excel file path
            include_text: Whether to include text in a separate sheet
            text_data: Text data by page number
            add_styling: Whether to add styling
            
        Returns:
            Path to created Excel file
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        sheet_count = 0
        
        # Add tables
        for page_num in sorted(tables.keys()):
            page_tables = tables[page_num]
            
            for table_idx, df in enumerate(page_tables, start=1):
                # Create sheet name
                if len(page_tables) == 1:
                    sheet_name = f"Page_{page_num}"
                else:
                    sheet_name = f"Page_{page_num}_Table_{table_idx}"
                
                sheet_name = sanitize_sheet_name(sheet_name)
                
                # Check sheet limit
                if sheet_count >= settings.EXCEL_MAX_SHEETS:
                    break
                
                # Create sheet and add data
                ws = wb.create_sheet(title=sheet_name)
                self._write_dataframe_to_sheet(ws, df, add_styling=add_styling)
                
                sheet_count += 1
        
        # Add text sheet if requested
        if include_text and text_data:
            text_sheet = wb.create_sheet(title="Text_Content")
            self._write_text_to_sheet(text_sheet, text_data, add_styling=add_styling)
        
        # If no sheets were added, add a placeholder
        if len(wb.sheetnames) == 0:
            ws = wb.create_sheet(title="No_Data")
            ws['A1'] = "No tables found in the PDF"
        
        # Save workbook
        try:
            wb.save(output_path)
        except Exception as e:
            raise ExportError(f"Failed to save Excel file: {str(e)}")
        
        return str(output_path)
    
    def export_tables_to_csv(
        self,
        tables: Dict[int, List[pd.DataFrame]],
        output_dir: str | Path,
        prefix: str = "table"
    ) -> List[str]:
        """
        Export tables to individual CSV files
        
        Args:
            tables: Dictionary mapping page numbers to list of DataFrames
            output_dir: Output directory for CSV files
            prefix: Filename prefix
            
        Returns:
            List of created CSV file paths
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        csv_files = []
        
        for page_num in sorted(tables.keys()):
            page_tables = tables[page_num]
            
            for table_idx, df in enumerate(page_tables, start=1):
                # Create filename
                if len(page_tables) == 1:
                    filename = f"{prefix}_page_{page_num}.csv"
                else:
                    filename = f"{prefix}_page_{page_num}_table_{table_idx}.csv"
                
                filepath = output_dir / filename
                
                # Export to CSV
                try:
                    df.to_csv(
                        filepath,
                        index=False,
                        encoding=settings.CSV_ENCODING,
                        sep=settings.CSV_DELIMITER
                    )
                    csv_files.append(str(filepath))
                except Exception as e:
                    raise ExportError(f"Failed to save CSV file {filename}: {str(e)}")
        
        return csv_files
    
    def create_combined_output(
        self,
        metadata: Dict[str, Any],
        text_data: Optional[Dict[int, str]] = None,
        tables: Optional[Dict[int, List[pd.DataFrame]]] = None
    ) -> Dict[str, Any]:
        """
        Create combined output with all data
        
        Args:
            metadata: PDF metadata
            text_data: Text data by page number
            tables: Tables by page number
            
        Returns:
            Combined dictionary with all data
        """
        output = {
            "metadata": metadata,
            "pages": []
        }
        
        # Determine page range
        all_pages = set()
        if text_data:
            all_pages.update(text_data.keys())
        if tables:
            all_pages.update(tables.keys())
        
        # If no data, use page count from metadata
        if not all_pages and metadata.get('page_count'):
            all_pages = set(range(1, metadata['page_count'] + 1))
        
        # Build page-by-page data
        for page_num in sorted(all_pages):
            page_data = {
                "page_number": page_num,
            }
            
            # Add text if available
            if text_data and page_num in text_data:
                page_data["text"] = text_data[page_num]
                page_data["char_count"] = len(text_data[page_num])
            
            # Add tables if available
            if tables and page_num in tables:
                page_tables = []
                for df in tables[page_num]:
                    page_tables.append({
                        "headers": df.columns.tolist(),
                        "rows": df.values.tolist(),
                        "row_count": len(df),
                        "col_count": len(df.columns)
                    })
                page_data["tables"] = page_tables
                page_data["table_count"] = len(page_tables)
            
            output["pages"].append(page_data)
        
        # Add statistics
        total_chars = sum(len(text) for text in (text_data or {}).values())
        total_tables = sum(len(page_tables) for page_tables in (tables or {}).values())
        
        output["statistics"] = {
            "total_pages": len(all_pages),
            "total_tables": total_tables,
            "total_chars": total_chars,
            "pages_with_tables": sorted(tables.keys()) if tables else []
        }
        
        # Add full text
        if text_data:
            output["all_text"] = "\n\n".join(
                text_data[page] for page in sorted(text_data.keys())
            )
        
        return output
    
    def _write_dataframe_to_sheet(
        self,
        worksheet,
        df: pd.DataFrame,
        add_styling: bool = True
    ) -> None:
        """
        Write DataFrame to Excel worksheet
        
        Args:
            worksheet: openpyxl worksheet
            df: DataFrame to write
            add_styling: Whether to add styling
        """
        # Check if table has real headers
        has_header = df.attrs.get('has_header', True)
        
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=has_header), start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
                
                # Apply styling to header row (only if table has real headers)
                if add_styling and has_header and r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif add_styling and not has_header and r_idx == 1:
                    # Add a note in the first row if no header
                    # Actually, let's add a comment to indicate no header
                    pass
        
        # Add note if no header detected
        if not has_header:
            # Add a note cell above the table
            worksheet.insert_rows(1)
            note_cell = worksheet.cell(row=1, column=1, value="Note: This table has no header row")
            note_cell.font = Font(italic=True, color="FF6B6B")
            # Merge cells for the note
            if len(df.columns) > 1:
                worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        
        # Auto-adjust column widths
        for col_idx, column in enumerate(worksheet.columns, start=1):
            max_length = 0
            
            for cell in column:
                # Skip merged cells
                if isinstance(cell, MergedCell):
                    continue
                    
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set column width
            if max_length > 0:
                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                column_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _write_text_to_sheet(
        self,
        worksheet,
        text_data: Dict[int, str],
        add_styling: bool = True
    ) -> None:
        """
        Write text data to Excel worksheet
        
        Args:
            worksheet: openpyxl worksheet
            text_data: Text data by page number
            add_styling: Whether to add styling
        """
        # Write headers
        worksheet['A1'] = "Page"
        worksheet['B1'] = "Text Content"
        
        if add_styling:
            for cell in ['A1', 'B1']:
                worksheet[cell].font = Font(bold=True, color="FFFFFF")
                worksheet[cell].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                worksheet[cell].alignment = Alignment(horizontal="center", vertical="center")
        
        # Write text data
        for row_idx, page_num in enumerate(sorted(text_data.keys()), start=2):
            worksheet.cell(row=row_idx, column=1, value=page_num)
            worksheet.cell(row=row_idx, column=2, value=text_data[page_num])
        
        # Adjust column widths
        worksheet.column_dimensions['A'].width = 10
        worksheet.column_dimensions['B'].width = 100

